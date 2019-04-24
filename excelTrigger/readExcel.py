import datetime
from openpyxl.reader.excel import load_workbook
from datetime import datetime
import xlrd
import os
import re

# version 2.6.2

# path = "C:\\Users\\Admin\\Desktop\\demo.xlsx"
'''
openpyxl.utils.exceptions.InvalidFileException: openpyxl does not support the old .xls file format, please use xlrd to read this file, or convert it to the more recent .xlsx file format.

'''


class readExcel(object):
    def __init__(self, filepath, filename,sheetname=None):
        self.filepath = filepath
        self.filename = filename
        self.fileType = None
        self.sheetname = sheetname
        self.excelDict = dict()

    def typeOfExcel(self):
        judgeExcel = {
            'xlsx': self.readXlsx,
            'xls': self.readXls,
        }
        self.fileType = re.split('\.', self.filename)[-1]
        # print(self.fileType)
        func = judgeExcel.get(self.fileType, 'error')
        return func()

    def readXls(self):
        xldata = xlrd.open_workbook(self.filepath)
        xltable = xldata.sheets()[0]
        nrows = xltable.nrows
        ncols = xltable.ncols
        countnum = 0
        for row in range(nrows):
            self.excelDict[countnum]=tuple(xltable.row_values(row))
            countnum = countnum + 1
        return self.excelDict

    def readXlsx(self):
        # print(self.filepath)
        workBooktObj = load_workbook(self.filepath)
        sheetNames = workBooktObj.get_sheet_names()
        sheetObj = workBooktObj.get_sheet_by_name(sheetNames[0])
        maxCol = sheetObj.max_column
        maxRow = sheetObj.max_row
        countnum = 0
        # Loop will print all columns name
        for items in sheetObj.values:
            self.excelDict[countnum]=items
            countnum=countnum+1
        # for row in range(1, maxRow + 1):
        #     for col in range(1, maxCol + 1):
        #         self.excelDict[row][col] = sheetObj.cell(row=row, column=col).value
                # cell_obj = sheetObj.cell(row=row, column=col)
                # print(cell_obj.value)
        return self.excelDict


if __name__ == '__main__':
    pass
    # path = os.path.abspath('移动直播2019年.xls')
    # print(path)
    # filename = '移动直播2019年.xls'
    # obj = readExcel(path, filename)
    # testdict = obj.typeOfExcel()
    # print(testdict)

    # print(re.split('/','2019/04/10'))
    # print(re.split(':', '18:04'))
